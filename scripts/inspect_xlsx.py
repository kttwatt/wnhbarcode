import sys
import openpyxl

path = r"h:\My Drive\วัสดุคงคลัง\พัสุดในคลังหน่วยเบิกกล่มงานการพยาบาลผู้ป่วยห้องผ่าตัด (1).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n===== SHEET:", ws.title, "dims:", ws.dimensions, "max_row:", ws.max_row, "max_col:", ws.max_column)
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # show all rows but trim very long
        vals = ["" if v is None else str(v) for v in row]
        print(i, "|", " | ".join(vals))
        if i >= 60:
            print("... (truncated)")
            break
